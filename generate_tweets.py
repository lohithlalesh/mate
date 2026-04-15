#!/usr/bin/env python3
"""
Simplify Suite — Tweet Generator
Fetches SaaS/business news, filters for relevance, generates brand-voice tweets.
Saves results to tweets_queue.xlsx in the repo.
"""

import os
import json
import time
import feedparser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from groq import Groq

# ── Configuration ─────────────────────────────────────────────────────────────

GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
OUTPUT_FILE    = "tweets_queue.xlsx"
MAX_PER_FEED   = 8      # articles pulled per RSS feed per run
MAX_TWEETS     = 12     # max tweets generated per run

RSS_FEEDS = [
    ("TechCrunch",   "https://techcrunch.com/feed/"),
    ("SaaStr",       "https://www.saastr.com/feed/"),
    ("G2 Learn",     "https://learn.g2.com/rss.xml"),
    ("The Verge",    "https://www.theverge.com/rss/index.xml"),
    ("Entrepreneur", "https://www.entrepreneur.com/latest.rss"),
    ("Product Hunt", "https://www.producthunt.com/feed"),
    ("HackerNews",   "https://news.ycombinator.com/rss"),
]

# ── Prompts ───────────────────────────────────────────────────────────────────

FILTER_PROMPT = """You are a content strategist for Simplify Suite, an all-in-one business management SaaS for small teams (1-25 people).

Review these news article titles and summaries. Return ONLY the ones worth tweeting about.

RELEVANT topics:
- SaaS tools, business software, project management, team productivity
- Small business / startup operations
- Expense management, cashflow, budgeting tools
- AI applied to business workflows
- Competitor news: monday.com, ClickUp, Asana, Notion, Rippling, Zoho, Freshworks, Trello, Odoo
- SaaS pricing, consolidation, tool fatigue, subscription costs
- Remote/hybrid work tools and operations

NOT RELEVANT:
- Consumer tech (phones, games, hardware)
- Pure politics, celebrity, sports
- Enterprise M&A with no SaaS angle

Return a JSON array only — no markdown, no explanation:
[{"index": 0, "reason": "one line why relevant"}]

If nothing is relevant return: []"""


TWEET_PROMPT = """You are the social media manager for Simplify Suite (@SimplifySuite).

Simplify Suite is an all-in-one business management platform built for small teams (1-25 people).
It replaces 5-10 separate SaaS tools: projects, budgets, cashflow, expenses, meetings, docs, performance reviews.
Flat pricing. No enterprise bloat. Built by a startup for startups.
Core message: "Stop paying for 5-10 SaaS subscriptions. One platform. Everything you need."

VOICE: Direct, sharp, slightly irreverent. Like a smart founder giving honest takes — not a brand account.

Given this news article, write ONE tweet that:
1. Reacts to or builds on the article in a way that's interesting to SMB founders and ops people
2. Uses one of these angles: hot take / pain point observation / stat + opinion / contrast / soft product spotlight / question
3. Feels like a real person wrote it
4. Is under 260 characters
5. Has 1-2 hashtags max — only if they feel natural, never forced
6. Occasionally (not every tweet) ties back to Simplify Suite's value

DO NOT:
- Start with "Hey" / "Breaking" / "Exciting news"
- Use more than one exclamation mark
- Use "synergy", "innovation", "game-changing", "disruptive"
- Be preachy
- Put the tweet in quotation marks

Return ONLY the tweet text. Nothing else."""


# ── Groq Setup ────────────────────────────────────────────────────────────────

def get_model():
    return Groq(api_key=GROQ_API_KEY)


def call_groq(client, system_prompt, user_content, retries=3):
    """Call Groq with simple retry logic."""
    for attempt in range(retries):
        try:
            response = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user",   "content": user_content},
                ],
                temperature=0.7,
                max_tokens=500,
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            print(f"  Groq error (attempt {attempt+1}): {e}")
            if attempt < retries - 1:
                time.sleep(5)
    return None


# ── RSS Fetching ──────────────────────────────────────────────────────────────

def fetch_all_articles():
    """Fetch articles from all RSS feeds."""
    articles = []
    for source_name, feed_url in RSS_FEEDS:
        try:
            feed = feedparser.parse(feed_url)
            count = 0
            for entry in feed.entries:
                if count >= MAX_PER_FEED:
                    break
                title = entry.get("title", "").strip()
                link  = entry.get("link", "").strip()
                # Get summary, strip HTML tags roughly
                summary = entry.get("summary", entry.get("description", ""))
                summary = summary.replace("<p>", " ").replace("</p>", " ")
                # Strip remaining HTML
                import re
                summary = re.sub(r"<[^>]+>", "", summary).strip()[:400]

                if title and link:
                    articles.append({
                        "title":   title,
                        "summary": summary,
                        "url":     link,
                        "source":  source_name,
                    })
                    count += 1
            print(f"  {source_name}: {count} articles fetched")
        except Exception as e:
            print(f"  {source_name}: failed — {e}")

    print(f"\nTotal articles fetched: {len(articles)}")
    return articles


# ── Deduplication ─────────────────────────────────────────────────────────────

def load_processed_urls():
    """Load URLs already in the XLSX so we don't regenerate tweets."""
    processed = set()
    if not os.path.exists(OUTPUT_FILE):
        return processed
    try:
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[3]:  # column D = Source URL
                processed.add(str(row[3]).strip())
    except Exception:
        pass
    return processed


def deduplicate(articles, processed_urls):
    """Remove articles already processed in previous runs."""
    fresh = [a for a in articles if a["url"] not in processed_urls]
    print(f"After dedup: {len(fresh)} new articles (skipped {len(articles) - len(fresh)} already processed)")
    return fresh


# ── Relevance Filtering ───────────────────────────────────────────────────────

def filter_relevant(client, articles):
    """Ask Groq to filter articles for relevance. Returns list of relevant articles."""
    if not articles:
        return []

    article_list = "\n".join(
        f'{i}. Title: "{a["title"]}" | Source: {a["source"]} | Summary: {a["summary"][:200]}'
        for i, a in enumerate(articles)
    )

    print("\nFiltering for relevance...")
    raw = call_groq(client, FILTER_PROMPT, f"Articles:\n{article_list}")

    if not raw:
        print("  Groq filter failed — skipping run")
        return []

    try:
        clean = raw.replace("```json", "").replace("```", "").strip()
        result = json.loads(clean)
        relevant_indices = [r["index"] for r in result if isinstance(r, dict) and "index" in r]
        relevant = [articles[i] for i in relevant_indices if i < len(articles)]
        print(f"  Relevant articles: {len(relevant)} of {len(articles)}")
        return relevant
    except json.JSONDecodeError:
        print(f"  Could not parse filter response: {raw[:200]}")
        return articles[:5]


# ── Tweet Generation ──────────────────────────────────────────────────────────

def generate_tweet(client, article):
    """Generate a single tweet for an article."""
    user_content = f"Article:\nTitle: {article['title']}\nSource: {article['source']}\nSummary: {article['summary']}"

    raw = call_groq(client, TWEET_PROMPT, user_content)
    if not raw:
        return None

    tweet = raw.strip().strip('"').strip("'")
    if tweet.lower().startswith("tweet:"):
        tweet = tweet[6:].strip()

    if len(tweet) < 20 or len(tweet) > 300:
        return None

    return tweet[:280]


# ── XLSX Output ───────────────────────────────────────────────────────────────

HEADERS = [
    "Tweet Text",
    "Source",
    "Article Title",
    "Source URL",
    "Generated At",
    "Status",       # leave blank; fill 'posted' when done
    "Notes",        # your personal notes
]

COL_WIDTHS = [80, 15, 50, 60, 20, 12, 30]

HEADER_FILL  = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL     = PatternFill("solid", fgColor="F5F7FF")
BORDER_SIDE  = Side(style="thin", color="CCCCCC")
CELL_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)


def create_or_load_workbook():
    """Load existing XLSX or create a new styled one."""
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        return wb, ws

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tweet Queue"

    # Write styled headers
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    return wb, ws


def append_tweets(wb, ws, new_rows):
    """Append new tweet rows with alternating row colors."""
    next_row = ws.max_row + 1

    for i, row_data in enumerate(new_rows):
        fill = ALT_FILL if (next_row + i) % 2 == 0 else None
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row + i, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = CELL_BORDER
            if fill:
                cell.fill = fill
        ws.row_dimensions[next_row + i].height = 60

    wb.save(OUTPUT_FILE)
    print(f"\nSaved {len(new_rows)} tweets to {OUTPUT_FILE}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not GROQ_API_KEY:
        raise RuntimeError("GROQ_API_KEY environment variable not set")

    print("=" * 60)
    print(f"Simplify Suite Tweet Generator — {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}")
    print("=" * 60)

    client = get_model()

    # 1. Fetch articles
    print("\n[1/4] Fetching RSS feeds...")
    articles = fetch_all_articles()

    # 2. Deduplicate against already-processed articles
    print("\n[2/4] Checking for already-processed articles...")
    processed_urls = load_processed_urls()
    fresh_articles = deduplicate(articles, processed_urls)

    if not fresh_articles:
        print("Nothing new to process. Exiting.")
        return

    # 3. Filter for relevance
    print("\n[3/4] Filtering for Simplify-relevant content...")
    relevant = filter_relevant(client, fresh_articles)

    if not relevant:
        print("No relevant articles found this run. Exiting.")
        return

    # Cap at max tweets per run
    relevant = relevant[:MAX_TWEETS]

    # 4. Generate tweets
    print(f"\n[4/4] Generating {len(relevant)} tweets...")
    new_rows = []
    for i, article in enumerate(relevant, start=1):
        print(f"  [{i}/{len(relevant)}] {article['title'][:60]}...")
        tweet = generate_tweet(client, article)
        if tweet:
            new_rows.append([
                tweet,
                article["source"],
                article["title"],
                article["url"],
                datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
                "",   # Status — blank until you post it
                "",   # Notes — for your use
            ])
            print(f"    → {tweet[:80]}...")
        else:
            print(f"    → skipped (bad response)")
        time.sleep(1)  # Be gentle with the API

    # 5. Save to XLSX
    wb, ws = create_or_load_workbook()
    if new_rows:
        append_tweets(wb, ws, new_rows)
    else:
        print("No valid tweets generated.")

    print("\nDone.")


if __name__ == "__main__":
    main()
